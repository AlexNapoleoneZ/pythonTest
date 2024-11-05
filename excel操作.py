import openpyxl

# 获取表单对象
wb = openpyxl.load_workbook('itemid.xlsx')
# ws = wb[wb.sheetnames[0]]
ws= wb.active

#打印坐标和值
# print(ws['A1'].coordinate)
# print(ws['A1'].value)
# print(ws['A1'].row)
# print(ws['A1'].column)
# print(ws.cell(row=1,column =2))
# print(ws.cell(row=1,column =2).value)

#获取行列
# colA = ws['A']
# for i in range(len(colA)):
#     print(colA[i].value)
# row1 = ws['1']
# for i in range(len(row1)):
#     print(row1[i].value)

# a = ''
# for i in range(1,9):
#     a = a + str(i)
# print(a)

rowTitle = ws['1']
annotation = ws['2']
#写入前半{}
print('local module = {}')
#写入注释
annotationText = '\n-- module["注释"] = {'
for cell in annotation:
    annotationText = annotationText + str(rowTitle[cell.column-1].value) + '=' + '"' + str(cell.value) + '"' + ','
annotationText = annotationText[ :-1]
annotationText = annotationText + '}\n'
print(annotationText)
#写入主体内容
for row in ws.iter_rows(min_row = 3):
    text = 'module[' + '"' + str(row[0].value) + '"' + '] = {'
    for cell in row:
        if cell.value != None:
            if type(cell.value) == type(rowTitle[cell.column-1].value):
                text = text + str(rowTitle[cell.column-1].value) + '=' + '"' + str(cell.value) + '"' + ','
            else:
                text = text + str(rowTitle[cell.column-1].value) + '=' + str(cell.value) + ','
    text = text[ :-1]
    text = text + '}'
    print(text)
#写入后半{}
print('return module')
